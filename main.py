# ===============================
# INICIO DEL PROGRAMA PRINCIPAL
# ===============================
# Este archivo es el punto de entrada de la aplicación de escritorio CNT-91.
# Al ejecutarlo, se carga la ventana principal con el diseño profesional definido en frontend_widgets.py

import tkinter as tk
from tkinter import ttk
from frontend_widgets import crear_layout_principal, get_info_cnt91_sections, get_info_cnt91_resources
import CNT_9X_pendulum as CNT
import os
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import numpy as np


# Variable global para el objeto del frecuencímetro
cnt_device = None


# Valor por defecto del identificador GPIB
DEFAULT_GPIB = 'GPIB0::10::INSTR'

# Variable global para el canal seleccionado ('A' o 'B')
canal_seleccionado = 'A'

# Variable global para el intervalo de captura (en segundos)
intervalo_s = 0.2

# Variable global para el acoplamiento ('AC' o 'DC')
acoplamiento = 'AC'

# Variable global para la impedancia ('Max' o 'Min')
impedancia = 'Max'  # Nuevo parámetro: 'Max' para 1MΩ, 'Min' para 50Ω

# Variable global para la atenuación ('1' o '10')
atenuacion = '1'  # Nuevo parámetro: '1' para 1x, '10' para 10x

# Variable global para el nivel de trigger (None para automático, float para manual)
trigger_level = None  # Nuevo parámetro: None para automático, float para manual (-5 a 5)

# Variable global para la pendiente del trigger ('POS' o 'NEG')
trigger_slope = 'POS'  # Nuevo parámetro: 'POS' para pendiente positiva, 'NEG' para negativa

# Variable global para el filtro analógico ajustable (None para False, 'True' para True)
filtro_Analog_PASSAbaja = None  # Nuevo parámetro: None para False, 'True' para True

# Variable global para rastrear si se ha guardado la configuración
configuracion_guardada = False

# Variable global para rastrear si la medición está pausada
medicion_pausada = False

# Variable global para almacenar la ruta del archivo Excel
ruta_archivo_excel = None

# Variable global para el buffer de mediciones
buffer_frecs_global = []

# Variable global para el nombre personalizado del archivo
nombre_archivo_var = None

# Diccionario global para guardar la configuración
global allan_config
allan_config = {}

# Función para mostrar el menú de selección de canal
def mostrar_menu_canal(widgets):
    global canal_seleccionado, intervalo_s, acoplamiento, impedancia, atenuacion, trigger_level, trigger_slope, filtro_Analog_PASSAbaja, configuracion_guardada, medicion_pausada, ruta_archivo_excel, nombre_archivo_var
    frame_contenido = widgets['frame_contenido']
    
    # Resetear el estado de configuración guardada
    configuracion_guardada = False
    
    # Resetear el estado de medición pausada
    medicion_pausada = False
    
    # Resetear la ruta del archivo Excel
    ruta_archivo_excel = None
    
    # Resetear el nombre del archivo
    nombre_archivo_var = None
    
    # Limpiar el frame de contenido
    for widget in frame_contenido.winfo_children():
        widget.destroy()
    
    # Frame principal con padding
    main_frame = tk.Frame(frame_contenido, bg='#f6f7fa')
    main_frame.pack(fill='both', expand=True, padx=10, pady=5)
    
    # Título principal - Frequency Datalogger
    titulo_frame = tk.Frame(main_frame, bg='#f6f7fa')
    titulo_frame.pack(fill='x', pady=(0, 10))
    
    titulo = tk.Label(titulo_frame, text='Frequency Datalogger', 
                     font=('Segoe UI', 14, 'bold'), 
                     fg='#25364a', bg='#f6f7fa')
    titulo.pack(anchor='w')
    
    # Línea separadora
    separador = tk.Frame(titulo_frame, height=1, bg='#2980f2')
    separador.pack(fill='x', pady=(3, 0))
    
    # Frame para las pestañas
    tabs_frame = tk.Frame(main_frame, bg='white', relief='solid', bd=1)
    tabs_frame.pack(fill='both', expand=True)
    
    # Frame superior para los botones de pestañas
    tabs_buttons_frame = tk.Frame(tabs_frame, bg='#f8f9fa', height=40)
    tabs_buttons_frame.pack(fill='x')
    tabs_buttons_frame.pack_propagate(False)
    
    # Frame para el contenido de las pestañas
    tabs_content_frame = tk.Frame(tabs_frame, bg='white')
    tabs_content_frame.pack(fill='both', expand=True)
    
    # Variables para controlar las pestañas
    current_tab = tk.StringVar(value='config')
    
    # Función para cambiar entre pestañas
    def switch_tab(tab_name):
        current_tab.set(tab_name)
        # Ocultar todos los frames de contenido
        for widget in tabs_content_frame.winfo_children():
            widget.pack_forget()
        
        # Mostrar el frame correspondiente
        if tab_name == 'config':
            config_canvas.pack(fill='both', expand=True)
            btn_config_tab.config(bg='#2980f2', fg='white')
            btn_datalogger_tab.config(bg='#f8f9fa', fg='#25364a')
        else:
            datalogger_content_frame.pack(fill='both', expand=True)
            btn_config_tab.config(bg='#f8f9fa', fg='#25364a')
            btn_datalogger_tab.config(bg='#2980f2', fg='white')
    
    # Botones de pestañas
    btn_config_tab = tk.Button(tabs_buttons_frame, text='Configuración', 
                              command=lambda: switch_tab('config'),
                              font=('Segoe UI', 10, 'bold'),
                              bg='#2980f2', fg='white',
                              relief='flat', padx=20, pady=8,
                              cursor='hand2')
    btn_config_tab.pack(side='left', padx=(10, 2))
    
    btn_datalogger_tab = tk.Button(tabs_buttons_frame, text='Datalogger', 
                                  command=lambda: switch_tab('datalogger'),
                                  font=('Segoe UI', 10, 'bold'),
                                  bg='#f8f9fa', fg='#25364a',
                                  relief='flat', padx=20, pady=8,
                                  cursor='hand2')
    btn_datalogger_tab.pack(side='left', padx=(2, 10))
    
    # ===== PESTAÑA DE CONFIGURACIÓN =====
    config_canvas = tk.Canvas(tabs_content_frame, bg='white', highlightthickness=0, bd=0)
    config_scrollbar = ttk.Scrollbar(tabs_content_frame, orient="vertical", command=config_canvas.yview)
    config_centrador = tk.Frame(config_canvas, bg='white')
    config_scrollable_frame = tk.Frame(config_centrador, bg='white')
    config_scrollable_frame.pack(anchor='nw', fill='both', expand=True)

    # Centrar el contenido horizontalmente al redimensionar
    def resize_config_canvas(event):
        canvas_width = event.width
        config_centrador.config(width=canvas_width)
        config_canvas.itemconfig(config_window_id, width=canvas_width)
    config_canvas.bind('<Configure>', resize_config_canvas)

    config_centrador.pack(expand=True)
    config_window_id = config_canvas.create_window((0, 0), window=config_centrador, anchor="n")
    config_canvas.configure(yscrollcommand=config_scrollbar.set)

    config_scrollable_frame.bind(
        "<Configure>",
        lambda e: config_canvas.configure(scrollregion=config_canvas.bbox("all"))
    )

    # Scroll con mouse wheel
    def _on_mousewheel_config(event):
        config_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
    config_canvas.bind_all("<MouseWheel>", _on_mousewheel_config)

    def _on_leave_config(event):
        config_canvas.unbind_all("<MouseWheel>")
    def _on_enter_config(event):
        config_canvas.bind_all("<MouseWheel>", _on_mousewheel_config)
    config_canvas.bind('<Enter>', _on_enter_config)
    config_canvas.bind('<Leave>', _on_leave_config)

    # Empaquetar canvas y scrollbar
    config_canvas.pack(side="left", fill="both", expand=True, padx=(0, 5))
    config_scrollbar.pack(side="right", fill="y")

    # Título de configuración
    config_titulo = tk.Label(config_scrollable_frame, text='Configuración:', 
                            font=('Segoe UI', 10, 'bold'), 
                            fg='#25364a', bg='white')
    config_titulo.pack(anchor='w', padx=10, pady=(8, 5))
    
    # Variable interna para el selector de canal
    canal_var = tk.StringVar(value='A')
    
    # Variable interna para el intervalo
    intervalo_var = tk.StringVar(value=str(intervalo_s))
    
    # Variable interna para el acoplamiento
    acoplamiento_var = tk.StringVar(value=acoplamiento)
    
    # Variable interna para la impedancia
    impedancia_var = tk.StringVar(value=impedancia)
    
    # Variable interna para la atenuación
    atenuacion_var = tk.StringVar(value=atenuacion)
    
    # Variable interna para el trigger
    trigger_mode_var = tk.StringVar(value='automatico' if trigger_level is None else 'manual')
    trigger_value_var = tk.StringVar(value=str(trigger_level) if trigger_level is not None else '0')
    
    # Variable interna para la pendiente del trigger
    trigger_slope_var = tk.StringVar(value=trigger_slope)
    
    # Variable interna para el filtro analógico
    filtro_analog_var = tk.StringVar(value='True' if filtro_Analog_PASSAbaja == 'True' else 'False')
    
    # Frame para el selector de canal
    selector_frame = tk.Frame(config_scrollable_frame, bg='white')
    selector_frame.pack(fill='x', padx=10, pady=(0, 8))
    
    # Etiqueta del selector de canal
    label_canal = tk.Label(selector_frame, text='Canal de medición:', 
                          font=('Segoe UI', 8, 'bold'), 
                          fg='#6c757d', bg='white')
    label_canal.pack(anchor='w', pady=(0, 3))
    
    # Frame para los radio buttons de canal
    radio_frame = tk.Frame(selector_frame, bg='white')
    radio_frame.pack(anchor='w')
    
    # Radio buttons con mejor estilo
    radio_a = tk.Radiobutton(radio_frame, text='Canal A', 
                            variable=canal_var, value='A', 
                            font=('Segoe UI', 8), 
                            fg='#25364a', bg='white',
                            selectcolor='white',
                            activebackground='white',
                            activeforeground='#2980f2')
    radio_a.pack(side='left', padx=(0, 15))
    
    radio_b = tk.Radiobutton(radio_frame, text='Canal B', 
                            variable=canal_var, value='B', 
                            font=('Segoe UI', 8), 
                            fg='#25364a', bg='white',
                            selectcolor='white',
                            activebackground='white',
                            activeforeground='#2980f2')
    radio_b.pack(side='left')
    
    # Separador entre configuraciones
    separador_config = tk.Frame(config_scrollable_frame, height=1, bg='#e0e0e0')
    separador_config.pack(fill='x', padx=10, pady=5)
    
    # Frame para el selector de intervalo
    intervalo_frame = tk.Frame(config_scrollable_frame, bg='white')
    intervalo_frame.pack(fill='x', padx=10, pady=(0, 8))
    
    # Etiqueta del selector de intervalo
    label_intervalo = tk.Label(intervalo_frame, text='Intervalo de captura (segundos):', 
                              font=('Segoe UI', 8, 'bold'), 
                              fg='#6c757d', bg='white')
    label_intervalo.pack(anchor='w', pady=(0, 3))
    
    # Etiqueta informativa de límites y valor por defecto
    label_intervalo_info = tk.Label(intervalo_frame, text='(2E-8 a 1000 s, por defecto 0.2 s)', 
                                   font=('Segoe UI', 7), 
                                   fg='#6c757d', bg='white')
    label_intervalo_info.pack(anchor='w', pady=(0, 3))
    
    # Frame para el control de intervalo
    control_intervalo_frame = tk.Frame(intervalo_frame, bg='white')
    control_intervalo_frame.pack(anchor='w')
    
    # Función para validar entrada manual
    def validar_entrada_intervalo(*args):
        try:
            texto_actual = intervalo_var.get()
            if texto_actual == "" or texto_actual == "0":
                return  # Permitir campo vacío o solo "0"
            
            valor = float(texto_actual)
            
            # Solo validar si el valor está completo y es menor que el mínimo
            if valor < 2e-8 and valor != 0:
                intervalo_var.set("2.00e-08")
            elif valor > 1000:
                intervalo_var.set("1000.000")
        except ValueError:
            # Si no es un número válido, no hacer nada (permitir que el usuario siga escribiendo)
            pass
    
    # Vincular validación a cambios en la variable
    intervalo_var.trace('w', validar_entrada_intervalo)
    
    # Entry para el valor del intervalo
    entry_intervalo = tk.Entry(control_intervalo_frame, 
                              textvariable=intervalo_var,
                              font=('Segoe UI', 8),
                              width=12,
                              justify='left')
    entry_intervalo.pack(side='left')
    
    # Separador entre configuraciones
    separador_config2 = tk.Frame(config_scrollable_frame, height=1, bg='#e0e0e0')
    separador_config2.pack(fill='x', padx=10, pady=5)
    
    # Frame para el selector de acoplamiento
    acoplamiento_frame = tk.Frame(config_scrollable_frame, bg='white')
    acoplamiento_frame.pack(fill='x', padx=10, pady=(0, 8))
    
    # Etiqueta del selector de acoplamiento
    label_acoplamiento = tk.Label(acoplamiento_frame, text='Acoplamiento:', 
                                 font=('Segoe UI', 8, 'bold'), 
                                 fg='#6c757d', bg='white')
    label_acoplamiento.pack(anchor='w', pady=(0, 3))
    
    # Frame para los radio buttons de acoplamiento
    radio_acoplamiento_frame = tk.Frame(acoplamiento_frame, bg='white')
    radio_acoplamiento_frame.pack(anchor='w')
    
    # Radio buttons para acoplamiento
    radio_ac = tk.Radiobutton(radio_acoplamiento_frame, text='AC', 
                             variable=acoplamiento_var, value='AC', 
                             font=('Segoe UI', 8), 
                             fg='#25364a', bg='white',
                             selectcolor='white',
                             activebackground='white',
                             activeforeground='#2980f2')
    radio_ac.pack(side='left', padx=(0, 15))
    
    radio_dc = tk.Radiobutton(radio_acoplamiento_frame, text='DC', 
                             variable=acoplamiento_var, value='DC', 
                             font=('Segoe UI', 8), 
                             fg='#25364a', bg='white',
                             selectcolor='white',
                             activebackground='white',
                             activeforeground='#2980f2')
    radio_dc.pack(side='left')
    
    # Separador entre configuraciones
    separador_config3 = tk.Frame(config_scrollable_frame, height=1, bg='#e0e0e0')
    separador_config3.pack(fill='x', padx=10, pady=5)
    
    # Frame para el selector de impedancia
    impedancia_frame = tk.Frame(config_scrollable_frame, bg='white')
    impedancia_frame.pack(fill='x', padx=10, pady=(0, 8))
    
    # Etiqueta del selector de impedancia
    label_impedancia = tk.Label(impedancia_frame, text='Impedancia:', 
                                 font=('Segoe UI', 8, 'bold'), 
                                 fg='#6c757d', bg='white')
    label_impedancia.pack(anchor='w', pady=(0, 3))
    
    # Frame para los radio buttons de impedancia
    radio_impedancia_frame = tk.Frame(impedancia_frame, bg='white')
    radio_impedancia_frame.pack(anchor='w')
    
    # Radio buttons para impedancia
    radio_max = tk.Radiobutton(radio_impedancia_frame, text='1MΩ (Max)', 
                             variable=impedancia_var, value='Max', 
                             font=('Segoe UI', 8), 
                             fg='#25364a', bg='white',
                             selectcolor='white',
                             activebackground='white',
                             activeforeground='#2980f2')
    radio_max.pack(side='left', padx=(0, 15))
    
    radio_min = tk.Radiobutton(radio_impedancia_frame, text='50Ω (Min)', 
                             variable=impedancia_var, value='Min', 
                             font=('Segoe UI', 8), 
                             fg='#25364a', bg='white',
                             selectcolor='white',
                             activebackground='white',
                             activeforeground='#2980f2')
    radio_min.pack(side='left')
    
    # Separador entre configuraciones
    separador_config4 = tk.Frame(config_scrollable_frame, height=1, bg='#e0e0e0')
    separador_config4.pack(fill='x', padx=10, pady=5)
    
    # Frame para el selector de atenuación
    atenuacion_frame = tk.Frame(config_scrollable_frame, bg='white')
    atenuacion_frame.pack(fill='x', padx=10, pady=(0, 8))
    
    # Etiqueta del selector de atenuación
    label_atenuacion = tk.Label(atenuacion_frame, text='Atenuación:', 
                                 font=('Segoe UI', 8, 'bold'), 
                                 fg='#6c757d', bg='white')
    label_atenuacion.pack(anchor='w', pady=(0, 3))
    
    # Frame para los radio buttons de atenuación
    radio_atenuacion_frame = tk.Frame(atenuacion_frame, bg='white')
    radio_atenuacion_frame.pack(anchor='w')
    
    # Radio buttons para atenuación
    radio_1x = tk.Radiobutton(radio_atenuacion_frame, text='1x', 
                             variable=atenuacion_var, value='1', 
                             font=('Segoe UI', 8), 
                             fg='#25364a', bg='white',
                             selectcolor='white',
                             activebackground='white',
                             activeforeground='#2980f2')
    radio_1x.pack(side='left', padx=(0, 15))
    
    radio_10x = tk.Radiobutton(radio_atenuacion_frame, text='10x', 
                             variable=atenuacion_var, value='10', 
                             font=('Segoe UI', 8), 
                             fg='#25364a', bg='white',
                             selectcolor='white',
                             activebackground='white',
                             activeforeground='#2980f2')
    radio_10x.pack(side='left')
    
    # Separador entre configuraciones
    separador_config5 = tk.Frame(config_scrollable_frame, height=1, bg='#e0e0e0')
    separador_config5.pack(fill='x', padx=10, pady=5)
    
    # Frame para el selector de trigger
    trigger_frame = tk.Frame(config_scrollable_frame, bg='white')
    trigger_frame.pack(fill='x', padx=10, pady=(0, 8))
    
    # Etiqueta del selector de trigger
    label_trigger = tk.Label(trigger_frame, text='Nivel de Trigger:', 
                            font=('Segoe UI', 8, 'bold'), 
                            fg='#6c757d', bg='white')
    label_trigger.pack(anchor='w', pady=(0, 3))
    
    # Etiqueta explicativa de porcentajes automáticos
    def actualizar_explicacion_trigger(*args):
        canal_actual = canal_var.get()
        if canal_actual == 'A':
            porcentaje = '70%'
        else:
            porcentaje = '30%'
        explicacion_trigger.config(text=f"(Automático: {porcentaje} del rango)")
    
    explicacion_trigger = tk.Label(trigger_frame, text='(Automático: 70% del rango)', 
                                  font=('Segoe UI', 7), 
                                  fg='#6c757d', bg='white')
    explicacion_trigger.pack(anchor='w', pady=(0, 3))
    
    # Vincular actualización de explicación a cambios en el canal
    canal_var.trace('w', actualizar_explicacion_trigger)
    
    # Frame para los radio buttons de trigger
    radio_trigger_frame = tk.Frame(trigger_frame, bg='white')
    radio_trigger_frame.pack(anchor='w')
    
    # Radio buttons para trigger
    radio_auto = tk.Radiobutton(radio_trigger_frame, text='Automático', 
                               variable=trigger_mode_var, value='automatico', 
                               font=('Segoe UI', 8), 
                               fg='#25364a', bg='white',
                               selectcolor='white',
                               activebackground='white',
                               activeforeground='#2980f2')
    radio_auto.pack(side='left', padx=(0, 15))
    
    radio_manual = tk.Radiobutton(radio_trigger_frame, text='Manual', 
                                 variable=trigger_mode_var, value='manual', 
                                 font=('Segoe UI', 8), 
                                 fg='#25364a', bg='white',
                                 selectcolor='white',
                                 activebackground='white',
                                 activeforeground='#2980f2')
    radio_manual.pack(side='left')
    
    # Frame para el valor manual del trigger
    trigger_value_frame = tk.Frame(trigger_frame, bg='white')
    trigger_value_frame.pack(anchor='w', pady=(5, 0))
    
    # Etiqueta del valor manual
    label_trigger_value = tk.Label(trigger_value_frame, text='Valor (-5 a 5 V):', 
                                  font=('Segoe UI', 8), 
                                  fg='#6c757d', bg='white')
    label_trigger_value.pack(side='left', padx=(0, 5))
    
    # Entry para el valor manual del trigger
    entry_trigger = tk.Entry(trigger_value_frame, 
                            textvariable=trigger_value_var,
                            font=('Segoe UI', 8),
                            width=8,
                            justify='left')
    entry_trigger.pack(side='left')
    
    # Función para validar entrada del trigger
    def validar_entrada_trigger(*args):
        try:
            texto_actual = trigger_value_var.get()
            if texto_actual == "":
                return  # Permitir campo vacío
            
            valor = float(texto_actual)
            
            # Validar límites
            if valor < -5:
                trigger_value_var.set("-5.0")
            elif valor > 5:
                trigger_value_var.set("5.0")
        except ValueError:
            # Si no es un número válido, no hacer nada
            pass
    
    # Vincular validación a cambios en la variable
    trigger_value_var.trace('w', validar_entrada_trigger)
    
    # Separador entre configuraciones
    separador_config6 = tk.Frame(config_scrollable_frame, height=1, bg='#e0e0e0')
    separador_config6.pack(fill='x', padx=10, pady=5)
    
    # Frame para el selector de pendiente del trigger
    trigger_slope_frame = tk.Frame(config_scrollable_frame, bg='white')
    trigger_slope_frame.pack(fill='x', padx=10, pady=(0, 8))
    
    # Etiqueta del selector de pendiente del trigger
    label_trigger_slope = tk.Label(trigger_slope_frame, text='Pendiente del Trigger:', 
                                  font=('Segoe UI', 8, 'bold'), 
                                  fg='#6c757d', bg='white')
    label_trigger_slope.pack(anchor='w', pady=(0, 3))
    
    # Frame para los radio buttons de pendiente del trigger
    radio_trigger_slope_frame = tk.Frame(trigger_slope_frame, bg='white')
    radio_trigger_slope_frame.pack(anchor='w')
    
    # Radio buttons para pendiente del trigger
    radio_pos = tk.Radiobutton(radio_trigger_slope_frame, text='Positiva', 
                              variable=trigger_slope_var, value='POS', 
                              font=('Segoe UI', 8), 
                              fg='#25364a', bg='white',
                              selectcolor='white',
                              activebackground='white',
                              activeforeground='#2980f2')
    radio_pos.pack(side='left', padx=(0, 15))
    
    radio_neg = tk.Radiobutton(radio_trigger_slope_frame, text='Negativa', 
                              variable=trigger_slope_var, value='NEG', 
                              font=('Segoe UI', 8), 
                              fg='#25364a', bg='white',
                              selectcolor='white',
                              activebackground='white',
                              activeforeground='#2980f2')
    radio_neg.pack(side='left')
    
    # Separador entre configuraciones
    separador_config7 = tk.Frame(config_scrollable_frame, height=1, bg='#e0e0e0')
    separador_config7.pack(fill='x', padx=10, pady=5)
    
    # Frame para el selector de filtro analógico
    filtro_analog_frame = tk.Frame(config_scrollable_frame, bg='white')
    filtro_analog_frame.pack(fill='x', padx=10, pady=(0, 8))
    
    # Etiqueta del selector de filtro analógico
    label_filtro_analog = tk.Label(filtro_analog_frame, text='Filtro Analógico pasa bajas:', 
                                  font=('Segoe UI', 8, 'bold'), 
                                  fg='#6c757d', bg='white')
    label_filtro_analog.pack(anchor='w', pady=(0, 3))
    
    # Frame para los radio buttons de filtro analógico
    radio_filtro_analog_frame = tk.Frame(filtro_analog_frame, bg='white')
    radio_filtro_analog_frame.pack(anchor='w')
    
    # Radio buttons para filtro analógico
    radio_filtro_true = tk.Radiobutton(radio_filtro_analog_frame, text='True', 
                                      variable=filtro_analog_var, value='True', 
                                      font=('Segoe UI', 8), 
                                      fg='#25364a', bg='white',
                                      selectcolor='white',
                                      activebackground='white',
                                      activeforeground='#2980f2')
    radio_filtro_true.pack(side='left', padx=(0, 15))
    
    radio_filtro_false = tk.Radiobutton(radio_filtro_analog_frame, text='False', 
                                       variable=filtro_analog_var, value='False', 
                                       font=('Segoe UI', 8), 
                                       fg='#25364a', bg='white',
                                       selectcolor='white',
                                       activebackground='white',
                                       activeforeground='#2980f2')
    radio_filtro_false.pack(side='left')
    
    # Función para guardar selección
    def guardar_seleccion():
        global canal_seleccionado, intervalo_s, acoplamiento, impedancia, atenuacion, trigger_level, trigger_slope, filtro_Analog_PASSAbaja, configuracion_guardada, ruta_archivo_excel, nombre_archivo_var
        canal_seleccionado = canal_var.get()
        acoplamiento = acoplamiento_var.get()
        impedancia = impedancia_var.get()
        atenuacion = atenuacion_var.get()
        trigger_slope = trigger_slope_var.get()
        
        # Procesar filtro analógico
        if filtro_analog_var.get() == 'True':
            filtro_Analog_PASSAbaja = 'True'
        else:
            filtro_Analog_PASSAbaja = None
        
        # Procesar trigger
        if trigger_mode_var.get() == 'automatico':
            trigger_level = None
        else:
            try:
                trigger_level = float(trigger_value_var.get())
                # Validar límites finales
                if trigger_level < -5:
                    trigger_level = -5.0
                elif trigger_level > 5:
                    trigger_level = 5.0
            except ValueError:
                trigger_level = 0.0  # Valor por defecto si hay error
        
        try:
            intervalo_s = float(intervalo_var.get())
            # Validar límites finales
            if intervalo_s < 2e-8:
                intervalo_s = 2e-8
            elif intervalo_s > 1000:
                intervalo_s = 1000
        except ValueError:
            intervalo_s = 0.2  # Valor por defecto si hay error
        
        # Formatear el intervalo para mostrar
        if intervalo_s < 0.001:
            intervalo_formato = f"{intervalo_s:.2e}"
        elif intervalo_s < 1:
            intervalo_formato = f"{intervalo_s:.6f}"
        else:
            intervalo_formato = f"{intervalo_s:.3f}"
        
        # Formatear trigger para mostrar
        if trigger_level is None:
            trigger_formato = "Automático"
        else:
            trigger_formato = f"{trigger_level:.1f}V"
        
        # Formatear filtro analógico para mostrar
        filtro_formato = "True" if filtro_Analog_PASSAbaja == 'True' else "False"
        
        resultado_label.config(text=f"✓ Configuración guardada: Canal {canal_seleccionado}, Intervalo {intervalo_formato} s, Acoplamiento {acoplamiento}, Impedancia {impedancia}, Atenuación {atenuacion}x, Trigger {trigger_formato}, Pendiente {trigger_slope}, Filtro Analógico pasa bajas {filtro_formato}", fg='#27ae60')
        
        # Convertir filtro_Analog_PASSAbaja a booleano real
        filtro_analog_bool = True if filtro_Analog_PASSAbaja == 'True' else False
        
        # Solo validar que el dispositivo esté conectado y guardar la configuración en memoria
        try:
            # Acceder al objeto CNT_91 global (cnt_device)
            if 'cnt_device' in globals() and cnt_device is not None:
                # Marcar configuración como guardada y habilitar botón de datalogger
                configuracion_guardada = True
                btn_start_stop.config(state='normal', bg='#27ae60', fg='white', cursor='hand2')
                btn_fin_medicion.config(state='normal')
                status_label.config(text='Estado: Listo para iniciar', fg='#27ae60')
                ruta_label.config(text='Archivo se creará al iniciar la medición', fg='#27ae60')
                
                # Habilitar el campo de nombre del archivo y limpiarlo
                entry_nombre_archivo.config(state='normal')
                nombre_archivo_var.set('')  # Limpiar el campo
                
            else:
                resultado_label.config(text="⚠ Error: Dispositivo no conectado", fg='#e74c3c')
        except Exception as e:
            resultado_label.config(text=f"⚠ Error al validar dispositivo: {str(e)}", fg='#e74c3c')
        
        # Actualizar la información de configuración en la pestaña de Datalogger
        actualizar_info_configuracion()
    
    # Frame para botón y resultado
    accion_frame = tk.Frame(config_scrollable_frame, bg='white')
    accion_frame.pack(fill='x', padx=10, pady=(0, 8))
    
    # Botón para guardar selección con estilo profesional
    btn_guardar = tk.Button(accion_frame, text='Guardar Configuración', 
                           command=guardar_seleccion, 
                           font=('Segoe UI', 8, 'bold'),
                           bg='#2980f2', fg='white',
                           relief='flat', padx=10, pady=3,
                           cursor='hand2')
    btn_guardar.pack(side='left', pady=(3, 0))
    
    # Label para mostrar resultado
    resultado_label = tk.Label(accion_frame, text='', 
                              font=('Segoe UI', 8), 
                              fg='#27ae60', bg='white')
    resultado_label.pack(side='left', padx=(8, 0), pady=(3, 0))
    
    # ===== PESTAÑA DE DATALOGGER =====
    datalogger_content_frame = tk.Frame(tabs_content_frame, bg='white')
    
    # Crear canvas y scrollbar para contenido scrolleable
    datalogger_canvas = tk.Canvas(datalogger_content_frame, bg='white', highlightthickness=0, bd=0)
    datalogger_scrollbar = ttk.Scrollbar(datalogger_content_frame, orient="vertical", command=datalogger_canvas.yview)
    
    # Frame centrador para el contenido
    datalogger_centrador = tk.Frame(datalogger_canvas, bg='white')
    datalogger_scrollable_frame = tk.Frame(datalogger_centrador, bg='white')
    datalogger_scrollable_frame.pack(anchor='center', expand=True)

    # Centrar el contenido horizontalmente al redimensionar
    def resize_datalogger_canvas(event):
        canvas_width = event.width
        datalogger_centrador.config(width=canvas_width)
        datalogger_canvas.itemconfig(datalogger_window_id, width=canvas_width)
    datalogger_canvas.bind('<Configure>', resize_datalogger_canvas)

    datalogger_centrador.pack(expand=True)
    datalogger_window_id = datalogger_canvas.create_window((0, 0), window=datalogger_centrador, anchor="n")
    datalogger_canvas.configure(yscrollcommand=datalogger_scrollbar.set)

    datalogger_scrollable_frame.bind(
        "<Configure>",
        lambda e: datalogger_canvas.configure(scrollregion=datalogger_canvas.bbox("all"))
    )
    
    # Configurar scroll con mouse wheel
    def _on_mousewheel(event):
        datalogger_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
    
    datalogger_canvas.bind_all("<MouseWheel>", _on_mousewheel)
    
    # Desvincular mouse wheel cuando se salga del canvas
    def _on_leave(event):
        datalogger_canvas.unbind_all("<MouseWheel>")
    
    def _on_enter(event):
        datalogger_canvas.bind_all("<MouseWheel>", _on_mousewheel)
    
    datalogger_canvas.bind('<Enter>', _on_enter)
    datalogger_canvas.bind('<Leave>', _on_leave)

    # Empaquetar canvas y scrollbar
    datalogger_canvas.pack(side="left", fill="both", expand=True, padx=(0, 5))
    datalogger_scrollbar.pack(side="right", fill="y")
    
    # Título de datalogger
    datalogger_titulo = tk.Label(datalogger_scrollable_frame, text='Datalogger:', 
                                font=('Segoe UI', 10, 'bold'), 
                                fg='#25364a', bg='white')
    datalogger_titulo.pack(anchor='w', padx=10, pady=(8, 5))
    
    # Frame para mostrar ruta del archivo Excel
    ruta_frame = tk.Frame(datalogger_scrollable_frame, bg='#e8f4fd', relief='solid', bd=1)
    ruta_frame.pack(fill='x', padx=10, pady=(0, 8))
    
    # Título de ruta
    ruta_titulo = tk.Label(ruta_frame, text='Archivo Excel:', 
                          font=('Segoe UI', 8, 'bold'), 
                          fg='#25364a', bg='#e8f4fd')
    ruta_titulo.pack(anchor='w', padx=10, pady=(5, 2))
    
    # Frame para nombre personalizado del archivo
    nombre_archivo_frame = tk.Frame(ruta_frame, bg='#e8f4fd')
    nombre_archivo_frame.pack(fill='x', padx=10, pady=(0, 5))
    
    # Etiqueta para el nombre del archivo
    label_nombre_archivo = tk.Label(nombre_archivo_frame, text='Nombre del archivo (opcional):', 
                                   font=('Segoe UI', 8), 
                                   fg='#6c757d', bg='#e8f4fd')
    label_nombre_archivo.pack(anchor='w', pady=(0, 2))
    
    # Etiqueta informativa
    label_info_nombre = tk.Label(nombre_archivo_frame, text='(Dejar vacío para nombre automático con fecha y hora)', 
                                font=('Segoe UI', 7), 
                                fg='#6c757d', bg='#e8f4fd')
    label_info_nombre.pack(anchor='w', pady=(0, 3))
    
    # Variable para el nombre del archivo
    nombre_archivo_var = tk.StringVar()
    
    # Entry para el nombre del archivo
    entry_nombre_archivo = tk.Entry(nombre_archivo_frame, 
                                   textvariable=nombre_archivo_var,
                                   font=('Segoe UI', 8),
                                   width=40,
                                   justify='left',
                                   state='disabled')  # Deshabilitado por defecto
    entry_nombre_archivo.pack(anchor='w', pady=(0, 5))
    
    # Label para mostrar ruta del archivo
    ruta_label = tk.Label(ruta_frame, text='No se ha configurado el dispositivo', 
                          font=('Segoe UI', 8), 
                          fg='#e74c3c', bg='#e8f4fd',
                          justify='left', anchor='nw')
    ruta_label.pack(anchor='w', padx=10, pady=(0, 5))
    
    # Frame para controles del datalogger
    datalogger_controls_frame = tk.Frame(datalogger_scrollable_frame, bg='white')
    datalogger_controls_frame.pack(fill='x', padx=10, pady=(0, 8))
    
    # Variable para el estado del datalogger
    datalogger_running = tk.BooleanVar(value=False)
    
    # Variables para almacenar datos de la gráfica
    frecuencias_grafica = []
    tiempos_relativos_grafica = []
    
    # Crear figura de matplotlib para la gráfica
    fig, ax = plt.subplots(figsize=(8, 4))
    ax.set_xlabel('Tiempo Relativo (s)', fontsize=10)
    ax.set_ylabel('Frecuencia (Hz)', fontsize=10)
    ax.set_title('Frecuencia vs Tiempo Relativo', fontsize=12, fontweight='bold')
    ax.grid(True, alpha=0.3)
    
    # Crear canvas de matplotlib
    canvas_frame = tk.Frame(datalogger_scrollable_frame, bg='white')
    canvas_frame.pack(fill='both', expand=True, padx=10, pady=(0, 10))
    
    canvas = FigureCanvasTkAgg(fig, canvas_frame)
    canvas_widget = canvas.get_tk_widget()
    canvas_widget.pack(fill='both', expand=True)
    
    # Frame para estadísticas
    stats_frame = tk.Frame(datalogger_scrollable_frame, bg='#f8f9fa', relief='solid', bd=1)
    stats_frame.pack(fill='x', padx=10, pady=(0, 10))
    
    # Título de estadísticas
    stats_titulo = tk.Label(stats_frame, text='Estadísticas en Tiempo Real:', 
                           font=('Segoe UI', 9, 'bold'), 
                           fg='#25364a', bg='#f8f9fa')
    stats_titulo.pack(anchor='w', padx=10, pady=(8, 5))
    
    # Label para mostrar estadísticas
    stats_label = tk.Label(stats_frame, text='', 
                          font=('Segoe UI', 8), 
                          fg='#2c3e50', bg='#f8f9fa',
                          justify='left', anchor='nw')
    stats_label.pack(anchor='w', padx=10, pady=(0, 8))
    
    # Función para actualizar la gráfica
    def actualizar_grafica():
        if len(frecuencias_grafica) > 0:
            # Limpiar gráfica anterior
            ax.clear()
            
            # Crear nueva gráfica
            ax.plot(tiempos_relativos_grafica, frecuencias_grafica, 'b-', linewidth=1, alpha=0.8)
            ax.scatter(tiempos_relativos_grafica, frecuencias_grafica, c='red', s=20, alpha=0.6)
            
            # Configurar ejes
            ax.set_xlabel('Tiempo Relativo (s)', fontsize=10)
            ax.set_ylabel('Frecuencia (Hz)', fontsize=10)
            ax.set_title('Frecuencia vs Tiempo Relativo', fontsize=12, fontweight='bold')
            ax.grid(True, alpha=0.3)
            
            # Auto-escalado inteligente
            if len(tiempos_relativos_grafica) > 1:
                # Para el eje X: mostrar los últimos 50 puntos o todos si hay menos
                x_max = max(tiempos_relativos_grafica)
                x_min = max(0, x_max - 50 * (x_max / len(tiempos_relativos_grafica)))
                ax.set_xlim(x_min, x_max)
                
                # Para el eje Y: margen del 5% arriba y abajo
                y_min, y_max = min(frecuencias_grafica), max(frecuencias_grafica)
                y_range = y_max - y_min
                if y_range > 0:
                    margin = y_range * 0.05
                    ax.set_ylim(y_min - margin, y_max + margin)
            
            # Actualizar canvas
            canvas.draw()
    
    # Función para actualizar estadísticas
    def actualizar_estadisticas():
        if len(frecuencias_grafica) > 0:
            freqs_array = np.array(frecuencias_grafica)
            
            # Calcular estadísticas
            maximo = np.max(freqs_array)
            minimo = np.min(freqs_array)
            media = np.mean(freqs_array)
            mediana = np.median(freqs_array)
            desv_tipica = np.std(freqs_array)
            varianza = np.var(freqs_array)
            
            # Formatear estadísticas
            stats_text = f"""Máximo: {maximo:.6f} Hz
Mínimo: {minimo:.6f} Hz
Media: {media:.6f} Hz
Mediana: {mediana:.6f} Hz
Desv. Típica: {desv_tipica:.6f} Hz
Varianza: {varianza:.6f} Hz²
Nº Muestras: {len(frecuencias_grafica)}"""
            
            stats_label.config(text=stats_text)
    
    # Función para agregar punto a la gráfica
    def agregar_punto_grafica(frecuencia, tiempo_relativo):
        frecuencias_grafica.append(frecuencia)
        tiempos_relativos_grafica.append(tiempo_relativo)
        
        # Actualizar gráfica y estadísticas
        actualizar_grafica()
        actualizar_estadisticas()
    
    # Función para finalizar medición definitivamente
    def finalizar_medicion():
        global medicion_pausada, buffer_frecs_global, nombre_archivo_var
        try:
            # Guardar cualquier medición pendiente en el buffer antes de finalizar
            if len(buffer_frecs_global) > 0:
                try:
                    for f, t, t_rel in buffer_frecs_global:
                        cnt_device.append_measurement(f, t, t_rel)
                    print(f"Mediciones pendientes guardadas al finalizar: {len(buffer_frecs_global)}")
                    buffer_frecs_global.clear()
                except Exception as e:
                    print(f"Advertencia al guardar mediciones pendientes al finalizar: {e}")
            
            # Guardar estadísticas finales en Excel si hay datos
            if len(frecuencias_grafica) > 0:
                try:
                    # Calcular estadísticas finales
                    freqs_array = np.array(frecuencias_grafica)
                    maximo = np.max(freqs_array)
                    minimo = np.min(freqs_array)
                    media = np.mean(freqs_array)
                    mediana = np.median(freqs_array)
                    desv_tipica = np.std(freqs_array)
                    varianza = np.var(freqs_array)
                    
                    # Guardar estadísticas en una nueva hoja del Excel
                    from openpyxl import load_workbook
                    wb = load_workbook(cnt_device.file_path)
                    
                    # Crear hoja de estadísticas
                    if 'Estadísticas' in wb.sheetnames:
                        ws_stats = wb['Estadísticas']
                    else:
                        ws_stats = wb.create_sheet('Estadísticas')
                    
                    # Limpiar hoja de estadísticas
                    ws_stats.delete_rows(1, ws_stats.max_row)
                    
                    # Escribir estadísticas
                    ws_stats.append(['ESTADÍSTICAS FINALES DE LA MEDICIÓN'])
                    ws_stats.append([])
                    ws_stats.append(['Parámetro', 'Valor', 'Unidad'])
                    ws_stats.append(['Máximo', maximo, 'Hz'])
                    ws_stats.append(['Mínimo', minimo, 'Hz'])
                    ws_stats.append(['Media', media, 'Hz'])
                    ws_stats.append(['Mediana', mediana, 'Hz'])
                    ws_stats.append(['Desviación Típica', desv_tipica, 'Hz'])
                    ws_stats.append(['Varianza', varianza, 'Hz²'])
                    ws_stats.append(['Número de Muestras', len(frecuencias_grafica), ''])
                    
                    # Guardar Excel
                    wb.save(cnt_device.file_path)
                    wb.close()
                    
                    print("Estadísticas finales guardadas en Excel.")
                    
                except Exception as e:
                    print(f"Advertencia al guardar estadísticas: {e}")
            
            # Cerrar el archivo Excel de manera definitiva
            try:
                cnt_device.cerrar_archivo_excel()
                print("Archivo Excel cerrado definitivamente.")
            except Exception as e:
                print(f"Advertencia al cerrar archivo Excel: {e}")
            
            # Abortar la medición continua si está activa
            try:
                cnt_device.abort_continuous_measurement()
            except Exception as e:
                print(f"Advertencia al abortar medición: {e}")
            
            # Resetear estados
            datalogger_running.set(False)
            medicion_pausada = False
            
            # Limpiar ruta del archivo Excel
            ruta_archivo_excel = None
            ruta_label.config(text='Archivo Excel cerrado', fg='#e74c3c')
            
            # Restaurar estado de la interfaz
            btn_start_stop.config(text='▶️  Iniciar Datalogger', bg='#27ae60', fg='white', cursor='hand2', state='normal')
            btn_fin_medicion.config(state='disabled')
            status_label.config(text='Estado: Medición finalizada', fg='#6c757d')
            
            # Habilitar el campo de nombre del archivo para nueva medición
            entry_nombre_archivo.config(state='normal')
            
            # Actualizar información de configuración
            actualizar_info_configuracion()
            
            tk.messagebox.showinfo('Medición Finalizada', 'La medición se ha finalizado definitivamente. Las estadísticas han sido guardadas en el archivo Excel.')
            
        except Exception as e:
            tk.messagebox.showerror('Error', f'Error al finalizar medición: {str(e)}')
    
    # Función para iniciar/detener datalogger
    def toggle_datalogger():
        global cnt_device, medicion_pausada, buffer_frecs_global, nombre_archivo_var
        
        if not datalogger_running.get():
            # Verificar que la configuración esté guardada
            if not configuracion_guardada:
                tk.messagebox.showerror('Error', 'Debe guardar la configuración antes de iniciar el datalogger.')
                return
            
            # Verificar que el dispositivo esté conectado
            if cnt_device is None:
                tk.messagebox.showerror('Error', 'Dispositivo no conectado. Conecte el CNT-91 primero.')
                return
            
            try:
                # Si es una nueva medición (no pausada), limpiar datos
                if not medicion_pausada:
                    # Limpiar datos de gráfica anteriores
                    frecuencias_grafica.clear()
                    tiempos_relativos_grafica.clear()
                    actualizar_grafica()
                    actualizar_estadisticas()
                    
                    # Configurar el dispositivo usando las variables globales
                    impedancia_convertida = 'MAX' if impedancia == 'Max' else 'MIN'
                    filtro_analog_bool = True if filtro_Analog_PASSAbaja == 'True' else False
                    
                    # Obtener nombre personalizado del archivo
                    nombre_personalizado = nombre_archivo_var.get().strip()
                    
                    # Si el usuario proporcionó un nombre, usarlo; si no, usar None para nombre automático
                    file_path_param = None
                    if nombre_personalizado:
                        # Asegurar que tenga extensión .xlsx
                        if not nombre_personalizado.endswith('.xlsx'):
                            nombre_personalizado += '.xlsx'
                        file_path_param = nombre_personalizado
                    
                    # Configurar el dispositivo
                    file_path = cnt_device.configurar_dispositivo(
                        canal=canal_seleccionado,
                        intervalo_s=intervalo_s,
                        acoplamiento=acoplamiento,
                        impedancia=impedancia_convertida,
                        atenuacion=atenuacion,
                        trigger_level=trigger_level,
                        trigger_slope=trigger_slope,
                        filtro_Digital_PASSAbaja=None,
                        filtro_Analog_PASSAbaja=filtro_analog_bool,
                        file_path=file_path_param
                    )
                    
                    # Actualizar la ruta del archivo Excel
                    ruta_archivo_excel = file_path
                    ruta_label.config(text=f'Guardando en: {file_path}', fg='#27ae60')
                    
                    # Iniciar medición continua
                    tiempo_espera = cnt_device.start_continuous_measurement(
                        intervalo_s=intervalo_s, 
                        canal=canal_seleccionado
                    )
                else:
                    # Si es reanudación, solo reiniciar la medición continua
                    tiempo_espera = cnt_device.start_continuous_measurement(
                        intervalo_s=intervalo_s, 
                        canal=canal_seleccionado
                    )
                
                # Calcular parámetros de buffer según el intervalo
                if intervalo_s < 2:
                    tiempo_espera = 0
                    lenght = 10
                elif intervalo_s < 5:
                    lenght = 1
                    tiempo_espera = 2.5 * (intervalo_s - 2) ** 2 + 0.09
                elif intervalo_s < 10:
                    lenght = 2
                    tiempo_espera = 1.2 * intervalo_s + 0.09
                else:
                    lenght = 1
                    tiempo_espera = intervalo_s + 5
                
                # Variables para el bucle de medición
                buffer_frecs = []
                buffer_frecs_global = buffer_frecs  # Hacer referencia global
                t0 = None  # Primer timestamp para calcular tiempos relativos
                
                # Si es reanudación, usar el último tiempo relativo como base
                if medicion_pausada and len(tiempos_relativos_grafica) > 0:
                    t0_offset = max(tiempos_relativos_grafica)
                else:
                    t0_offset = 0
                
                # Función interna para el bucle de medición
                def medicion_loop():
                    nonlocal buffer_frecs, t0
                    
                    if datalogger_running.get():
                        try:
                            # Obtener muestras
                            frecs, ts = cnt_device.fetch_continuous_samples(
                                n_muestras=1,
                                tiempo_espera=tiempo_espera
                            )
                            
                            for f, t in zip(frecs, ts):
                                if t0 is None:
                                    t0 = t
                                t_rel = t - t0 + t0_offset
                                buffer_frecs.append((f, t, t_rel))
                                
                                # Agregar punto a la gráfica en tiempo real
                                agregar_punto_grafica(f, t_rel)
                                
                                # Actualizar estado en la interfaz
                                status_label.config(text=f'Estado: Midiendo... Frecuencia: {f:.6f} Hz, T. Relativo: {t_rel:.3f} s', fg='#27ae60')
                                
                                # Actualizar información de configuración con datos en tiempo real
                                info_text = f"""Canal: {canal_seleccionado}
Intervalo: {intervalo_s:.6f} s
Acoplamiento: {acoplamiento}
Impedancia: {impedancia}
Atenuación: {atenuacion}x
Trigger: {'Automático' if trigger_level is None else f'{trigger_level:.1f}V'}
Pendiente: {trigger_slope}
Filtro Analógico: {'True' if filtro_Analog_PASSAbaja == 'True' else 'False'}

Última medición:
Frecuencia: {f:.6f} Hz
Timestamp: {t:.6f} s
T. Relativo: {t_rel:.6f} s"""
                                info_label.config(text=info_text)
                            
                            # Guardar en Excel cuando el buffer esté lleno
                            if len(buffer_frecs) >= lenght:
                                for f, t, t_rel in buffer_frecs:
                                    cnt_device.append_measurement(f, t, t_rel)
                                buffer_frecs.clear()
                                buffer_frecs_global.clear()  # Limpiar también la referencia global
                            
                            # Programar la siguiente medición
                            root.after(100, medicion_loop)  # 100ms entre mediciones
                            
                        except Exception as e:
                            tk.messagebox.showerror('Error', f'Error durante la medición: {str(e)}')
                            datalogger_running.set(False)
                            btn_start_stop.config(text='▶️  Iniciar Datalogger', bg='#27ae60', fg='white', cursor='hand2')
                            status_label.config(text='Estado: Error en medición', fg='#e74c3c')
                
                # Iniciar el bucle de medición
                datalogger_running.set(True)
                medicion_pausada = False
                btn_start_stop.config(text='⏸️  Pausar Datalogger', bg='#f39c12', fg='white', cursor='hand2')
                btn_fin_medicion.config(state='normal')
                status_label.config(text='Estado: Iniciando medición...', fg='#27ae60')
                
                # Deshabilitar el campo de nombre del archivo
                entry_nombre_archivo.config(state='disabled')
                
                # Iniciar el bucle de medición
                medicion_loop()
                
            except Exception as e:
                tk.messagebox.showerror('Error', f'Error al iniciar datalogger: {str(e)}')
                datalogger_running.set(False)
                btn_start_stop.config(text='▶️  Iniciar Datalogger', bg='#27ae60', fg='white', cursor='hand2')
                status_label.config(text='Estado: Error al iniciar', fg='#e74c3c')
        else:
            # Pausar datalogger (no cerrar Excel)
            try:
                datalogger_running.set(False)
                medicion_pausada = True
                
                # Guardar cualquier medición pendiente en el buffer antes de pausar
                if len(buffer_frecs_global) > 0:
                    try:
                        for f, t, t_rel in buffer_frecs_global:
                            cnt_device.append_measurement(f, t, t_rel)
                        print(f"Mediciones pendientes guardadas al pausar: {len(buffer_frecs_global)}")
                        buffer_frecs_global.clear()
                    except Exception as e:
                        print(f"Advertencia al guardar mediciones pendientes: {e}")
                
                # Solo abortar la medición continua, NO cerrar Excel
                try:
                    cnt_device.abort_continuous_measurement()
                except Exception as e:
                    print(f"Advertencia al abortar medición: {e}")
                
                # Cambiar botón a "Reanudar"
                btn_start_stop.config(text='▶️  Reanudar Datalogger', bg='#27ae60', fg='white', cursor='hand2')
                status_label.config(text='Estado: Datalogger pausado', fg='#f39c12')
                
                # Mantener el campo de nombre del archivo deshabilitado
                entry_nombre_archivo.config(state='disabled')
                
                tk.messagebox.showinfo('Datalogger Pausado', 'La medición se ha pausado. Puede reanudarla o finalizarla definitivamente.')
                
            except Exception as e:
                tk.messagebox.showerror('Error', f'Error al pausar datalogger: {str(e)}')
    
    # Botón para iniciar/detener datalogger
    btn_start_stop = tk.Button(datalogger_controls_frame, text='▶️  Iniciar Datalogger', 
                              command=toggle_datalogger, 
                              font=('Segoe UI', 10, 'bold'),
                              bg='#cccccc', fg='#666666',
                              relief='flat', padx=15, pady=5,
                              cursor='arrow',
                              state='disabled')
    btn_start_stop.pack(side='left', pady=(3, 0))
    
    # Botón para finalizar medición (rojo)
    btn_fin_medicion = tk.Button(datalogger_controls_frame, text='🔴  FIN DE MEDICIÓN', 
                                command=lambda: finalizar_medicion(),
                                font=('Segoe UI', 10, 'bold'),
                                bg='#e74c3c', fg='white',
                                relief='flat', padx=15, pady=5,
                                cursor='hand2',
                                state='disabled')
    btn_fin_medicion.pack(side='left', padx=(10, 0), pady=(3, 0))
    
    # Label para mostrar estado del datalogger
    status_label = tk.Label(datalogger_controls_frame, text='Estado: Configure el dispositivo primero', 
                           font=('Segoe UI', 9), 
                           fg='#e74c3c', bg='white')
    status_label.pack(side='left', padx=(15, 0), pady=(3, 0))
    
    # Separador
    separador_datalogger = tk.Frame(datalogger_scrollable_frame, height=1, bg='#e0e0e0')
    separador_datalogger.pack(fill='x', padx=10, pady=10)
    
    # Frame para información de configuración actual
    info_frame = tk.Frame(datalogger_scrollable_frame, bg='#f8f9fa', relief='solid', bd=1)
    info_frame.pack(fill='x', padx=10, pady=(0, 10))
    
    # Título de información
    info_titulo = tk.Label(info_frame, text='Configuración Actual:', 
                          font=('Segoe UI', 9, 'bold'), 
                          fg='#25364a', bg='#f8f9fa')
    info_titulo.pack(anchor='w', padx=10, pady=(8, 5))
    
    # Label para mostrar información de configuración
    info_label = tk.Label(info_frame, text='', 
                         font=('Segoe UI', 8), 
                         fg='#2c3e50', bg='#f8f9fa',
                         justify='left', anchor='nw')
    info_label.pack(anchor='w', padx=10, pady=(0, 8))
    
    # Función para actualizar información de configuración
    def actualizar_info_configuracion():
        # Formatear valores para mostrar
        if intervalo_s < 0.001:
            intervalo_formato = f"{intervalo_s:.2e}"
        elif intervalo_s < 1:
            intervalo_formato = f"{intervalo_s:.6f}"
        else:
            intervalo_formato = f"{intervalo_s:.3f}"
        
        trigger_formato = "Automático" if trigger_level is None else f"{trigger_level:.1f}V"
        filtro_formato = "True" if filtro_Analog_PASSAbaja == 'True' else "False"
        
        info_text = f"""Canal: {canal_seleccionado}
Intervalo: {intervalo_formato} s
Acoplamiento: {acoplamiento}
Impedancia: {impedancia}
Atenuación: {atenuacion}x
Trigger: {trigger_formato}
Pendiente: {trigger_slope}
Filtro Analógico: {filtro_formato}"""
        
        info_label.config(text=info_text)
    
    # Mostrar la pestaña de configuración por defecto
    switch_tab('config')

# Función para mostrar la página de Allan Deviation vs tau
def mostrar_allan_deviation(widgets):
    frame_contenido = widgets['frame_contenido']
    
    # Limpiar el frame de contenido
    for widget in frame_contenido.winfo_children():
        widget.destroy()
    
    # Frame principal con padding
    main_frame = tk.Frame(frame_contenido, bg='#f6f7fa')
    main_frame.pack(fill='both', expand=True, padx=10, pady=5)
    
    # Título principal - Allan Deviation vs tau
    titulo_frame = tk.Frame(main_frame, bg='#f6f7fa')
    titulo_frame.pack(fill='x', pady=(0, 10))
    
    titulo = tk.Label(titulo_frame, text='Allan Deviation vs tau', 
                     font=('Segoe UI', 16, 'bold'), 
                     fg='#25364a', bg='#f6f7fa')
    titulo.pack(anchor='w')
    
    # Línea separadora azul
    separador = tk.Frame(titulo_frame, height=2, bg='#2980f2')
    separador.pack(fill='x', pady=(5, 0))
    
    # Frame de contenido con scroll
    contenido_frame = tk.Frame(main_frame, bg='white', relief='solid', bd=1)
    contenido_frame.pack(fill='both', expand=True, pady=(0, 5))
    
    # Frame para los botones en la esquina superior izquierda (fuera del canvas)
    botones_frame = tk.Frame(contenido_frame, bg='white')
    botones_frame.pack(anchor='nw', padx=8, pady=(10, 0))
    
    # Botón Configuración
    btn_configuracion = tk.Button(botones_frame, text='Configuración', 
                                 font=('Segoe UI', 10, 'bold'),
                                 bg='#2980f2', fg='white',
                                 relief='flat', padx=15, pady=5,
                                 cursor='hand2')
    btn_configuracion.pack(side='left', padx=(0, 10))
    
    # Botón Allan Deviation
    btn_allan_dev = tk.Button(botones_frame, text='Allan Deviation', 
                             font=('Segoe UI', 10, 'bold'),
                             bg='#6c757d', fg='white',
                             relief='flat', padx=15, pady=5,
                             cursor='hand2')
    btn_allan_dev.pack(side='left')

    # Canvas y scrollbar para contenido scrolleable (debajo de los botones)
    canvas = tk.Canvas(contenido_frame, bg='white', highlightthickness=0, bd=0)
    scrollbar = ttk.Scrollbar(contenido_frame, orient="vertical", command=canvas.yview)

    centrador = tk.Frame(canvas, bg='white')
    scrollable_frame = tk.Frame(centrador, bg='white')
    scrollable_frame.pack(anchor='center', expand=True)

    def resize_canvas(event):
        canvas_width = event.width
        centrador.config(width=canvas_width)
        canvas.itemconfig(window_id, width=canvas_width)
    canvas.bind('<Configure>', resize_canvas)

    centrador.pack(expand=True)
    window_id = canvas.create_window((0, 0), window=centrador, anchor="n")
    canvas.configure(yscrollcommand=scrollbar.set)

    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )
    
    def _on_mousewheel(event):
        canvas.yview_scroll(int(-1*(event.delta/120)), "units")
    canvas.bind_all("<MouseWheel>", _on_mousewheel)
    def _on_leave(event):
        canvas.unbind_all("<MouseWheel>")
    def _on_enter(event):
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
    canvas.bind('<Enter>', _on_enter)
    canvas.bind('<Leave>', _on_leave)

    canvas.pack(side="left", fill="both", expand=True, padx=(0, 5), pady=10)
    scrollbar.pack(side="right", fill="y", padx=(0, 10), pady=10)
    scrollable_frame.pack_configure(padx=10)

    # ===== CONTENIDO DE ALLAN DEVIATION VS TAU =====
    # Crear frames para cada pestaña
    frame_config = tk.Frame(scrollable_frame, bg='white')
    frame_allan = tk.Frame(scrollable_frame, bg='white')

    # Variable para el canal seleccionado
    canal_allan_var = tk.StringVar(value='A')
    
    # ===== CONTENIDO DE LA PESTAÑA ALLAN DEVIATION =====
    # Título de Allan Deviation vs tau
    titulo_allan = tk.Label(frame_allan, text='Allan Deviation vs tau:', 
                           font=('Segoe UI', 12, 'bold'), 
                           fg='#25364a', bg='white')
    titulo_allan.pack(anchor='w', padx=20, pady=(20, 10))
    
    # Frame para el botón de medición
    medicion_frame = tk.Frame(frame_allan, bg='white')
    medicion_frame.pack(anchor='w', padx=20, pady=(0, 10))
    
    # Función para ejecutar la medición Allan Deviation
    def ejecutar_medicion_allan():
        try:
            # Verificar que el dispositivo esté conectado
            if 'cnt_device' not in globals() or cnt_device is None:
                tk.messagebox.showerror('Error', 'Dispositivo no conectado. Conecte el CNT-91 primero.')
                return
            
            # Verificar que la configuración esté guardada
            if not allan_config:
                tk.messagebox.showerror('Error', 'Debe guardar la configuración antes de ejecutar la medición.')
                return
            
            # Obtener el nombre del archivo
            nombre_archivo = nombre_archivo_allan_var.get().strip()
            if nombre_archivo == 'None' or nombre_archivo == '':
                nombre_archivo = None  # Usar nombre automático
            
            # Convertir parámetros de configuración al formato esperado por la función
            canal = allan_config['canal']
            N_muestras = allan_config['N_muestras']
            intervalo_captura_min = allan_config['intervalo_captura_min']
            intervalo_captura_max = allan_config['intervalo_captura_max']
            pasos = allan_config['pasos']
            acoplamiento = allan_config['acoplamiento']
            impedancia = allan_config['impedancia']
            atenuacion = allan_config['atenuacion']
            trigger_level = allan_config['trigger_level']
            trigger_slope = allan_config['trigger_slope']
            filtro_Analog_PASSAbaja = allan_config['filtro_Analog_PASSAbaja']
            
            # Mostrar mensaje de inicio
            tk.messagebox.showinfo('Iniciando Medición', f'Iniciando medición Allan Deviation con {N_muestras} muestras y {pasos} pasos...')
            
            # Ejecutar la medición
            resultados = cnt_device.calc_Adev_EstadisticsyADEV(
                canal=canal,
                N_muestras=N_muestras,
                intervalo_captura_min=intervalo_captura_min,
                intervalo_captura_max=intervalo_captura_max,
                pasos=pasos,
                acoplamiento=acoplamiento,
                impedancia=impedancia,
                atenuacion=atenuacion,
                trigger_level=trigger_level,
                trigger_slope=trigger_slope,
                filtro_Analog_PASSAbaja=filtro_Analog_PASSAbaja,
                guardar=True,
                nombre_archivo=nombre_archivo
            )
            
            # Mostrar resultados
            if resultados:
                # Crear ventana de resultados
                ventana_resultados = tk.Toplevel()
                ventana_resultados.title('Resultados Allan Deviation')
                ventana_resultados.geometry('800x600')
                ventana_resultados.configure(bg='white')
                
                # Frame principal
                main_frame = tk.Frame(ventana_resultados, bg='white')
                main_frame.pack(fill='both', expand=True, padx=10, pady=10)
                
                # Título
                titulo = tk.Label(main_frame, text='Resultados Allan Deviation', 
                                font=('Segoe UI', 14, 'bold'), 
                                fg='#25364a', bg='white')
                titulo.pack(pady=(0, 10))
                
                # Información general
                info_text = f"Se realizaron {len(resultados)} mediciones\n"
                info_text += f"Canal: {canal} | Muestras: {N_muestras} | Pasos: {pasos}\n"
                if nombre_archivo:
                    info_text += f"Archivo: {nombre_archivo}.xlsx"
                else:
                    # Obtener el nombre del archivo que realmente se generó
                    try:
                        # Buscar el archivo más reciente en la carpeta Mediciones_CNT
                        import os
                        import glob
                        from datetime import datetime
                        
                        carpeta = 'Mediciones_CNT'
                        if os.path.exists(carpeta):
                            # Buscar archivos .xlsx en la carpeta
                            archivos = glob.glob(os.path.join(carpeta, "*.xlsx"))
                            if archivos:
                                # Obtener el archivo más reciente
                                archivo_mas_reciente = max(archivos, key=os.path.getctime)
                                nombre_archivo_real = os.path.basename(archivo_mas_reciente)
                                info_text += f"Archivo: {carpeta}\\{nombre_archivo_real}"
                            else:
                                info_text += "Archivo: No encontrado"
                        else:
                            info_text += "Archivo: Carpeta no encontrada"
                    except Exception as e:
                        info_text += f"Archivo: Error al obtener nombre ({str(e)})"
                
                info_label = tk.Label(main_frame, text=info_text, 
                                    font=('Segoe UI', 10), 
                                    fg='#2c3e50', bg='white')
                info_label.pack(pady=(0, 10))
                
                # Crear tabla de resultados
                tabla_frame = tk.Frame(main_frame, bg='white')
                tabla_frame.pack(fill='both', expand=True)
                
                # Encabezados de la tabla
                headers = ['Número', 'Intervalo[s]', 'ADEV[Hz]', 'Estabilidad[-]']
                for i, header in enumerate(headers):
                    label = tk.Label(tabla_frame, text=header, 
                                   font=('Segoe UI', 9, 'bold'), 
                                   fg='#25364a', bg='#f8f9fa',
                                   relief='solid', bd=1)
                    label.grid(row=0, column=i, sticky='ew', padx=1, pady=1)
                
                # Datos de la tabla
                for i, resultado in enumerate(resultados, 1):
                    numero = resultado['Numero de medida']
                    intervalo = resultado['intervalo_captura']
                    adev = resultado['allan_deviation']
                    estabilidad = resultado['Estabilidad_Adev']
                    
                    # Formatear valores
                    if adev is not None:
                        adev_str = f"{adev:.3e}"
                    else:
                        adev_str = "N/A"
                    
                    if estabilidad is not None:
                        estabilidad_str = f"{estabilidad:.3e}"
                    else:
                        estabilidad_str = "N/A"
                    
                    # Crear filas de datos
                    tk.Label(tabla_frame, text=f"{numero}", 
                            font=('Segoe UI', 8), bg='white',
                            relief='solid', bd=1).grid(row=i, column=0, sticky='ew', padx=1, pady=1)
                    tk.Label(tabla_frame, text=f"{intervalo:.6f}", 
                            font=('Segoe UI', 8), bg='white',
                            relief='solid', bd=1).grid(row=i, column=1, sticky='ew', padx=1, pady=1)
                    tk.Label(tabla_frame, text=adev_str, 
                            font=('Segoe UI', 8), bg='white',
                            relief='solid', bd=1).grid(row=i, column=2, sticky='ew', padx=1, pady=1)
                    tk.Label(tabla_frame, text=estabilidad_str, 
                            font=('Segoe UI', 8), bg='white',
                            relief='solid', bd=1).grid(row=i, column=3, sticky='ew', padx=1, pady=1)
                
                # Configurar columnas para que se expandan
                tabla_frame.columnconfigure(0, weight=1)
                tabla_frame.columnconfigure(1, weight=1)
                tabla_frame.columnconfigure(2, weight=1)
                tabla_frame.columnconfigure(3, weight=1)
                
                # Crear gráficas
                try:
                    # Extraer datos para gráficas
                    intervalos = [r['intervalo_captura'] for r in resultados]
                    adevs = [r['allan_deviation'] for r in resultados if r['allan_deviation'] is not None]
                    estabilidades = [r['Estabilidad_Adev'] for r in resultados if r['Estabilidad_Adev'] is not None]
                    
                    if len(adevs) > 0 and len(estabilidades) > 0:
                        # Crear figura con matplotlib
                        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5))
                        
                        # Gráfica ADEV vs intervalo
                        ax1.loglog(intervalos[:len(adevs)], adevs, 'bo-', linewidth=2, markersize=8)
                        ax1.set_xlabel('Intervalo de captura τ [s]')
                        ax1.set_ylabel('Allan Deviation σ(τ) [Hz]')
                        ax1.set_title('ADEV vs Intervalo de Captura (Log-Log)')
                        ax1.grid(True, which='both', alpha=0.3)
                        
                        # Gráfica Estabilidad vs intervalo
                        ax2.loglog(intervalos[:len(estabilidades)], estabilidades, 'ro-', linewidth=2, markersize=8)
                        ax2.set_xlabel('Intervalo de captura τ [s]')
                        ax2.set_ylabel('Estabilidad ADEV σ(τ)/f₀ [-]')
                        ax2.set_title('Estabilidad ADEV vs Intervalo de Captura (Log-Log)')
                        ax2.grid(True, which='both', alpha=0.3)
                        
                        plt.tight_layout()
                        plt.show()
                        
                except Exception as e:
                    print(f"Error al crear gráficas: {e}")
                
                # Botón para cerrar
                btn_cerrar = tk.Button(main_frame, text='Cerrar', 
                                     command=ventana_resultados.destroy,
                                     font=('Segoe UI', 10, 'bold'),
                                     bg='#6c757d', fg='white',
                                     relief='flat', padx=20, pady=5)
                btn_cerrar.pack(pady=(10, 0))
                
            else:
                tk.messagebox.showerror('Error', 'No se obtuvieron resultados de la medición.')
                
        except Exception as e:
            tk.messagebox.showerror('Error', f'Error durante la medición: {str(e)}')
    
    # Botón Medición
    btn_medicion_allan = tk.Button(medicion_frame, text='Medición', 
                                  command=ejecutar_medicion_allan,
                                  font=('Segoe UI', 10, 'bold'),
                                  bg='#2980f2', fg='white',
                                  relief='flat', padx=15, pady=5,
                                  cursor='hand2')
    btn_medicion_allan.pack(anchor='w', pady=(0, 5))
    
    # Frame para el nombre del archivo
    nombre_archivo_allan_frame = tk.Frame(frame_allan, bg='white')
    nombre_archivo_allan_frame.pack(anchor='w', padx=20, pady=(0, 10))
    
    # Etiqueta para el nombre del archivo
    label_nombre_archivo_allan = tk.Label(nombre_archivo_allan_frame, text='Nombre archivo:', 
                                         font=('Segoe UI', 10, 'bold'), 
                                         fg='#25364a', bg='white')
    label_nombre_archivo_allan.pack(anchor='w', pady=(0, 3))
    
    # Variable para el nombre del archivo Allan Deviation
    nombre_archivo_allan_var = tk.StringVar(value='None')
    
    # Entry para el nombre del archivo
    entry_nombre_archivo_allan = tk.Entry(nombre_archivo_allan_frame, 
                                         textvariable=nombre_archivo_allan_var,
                                         font=('Segoe UI', 10),
                                         width=30,
                                         justify='left')
    entry_nombre_archivo_allan.pack(anchor='w', pady=(0, 5))

    # Frame para el selector de canal
    selector_frame = tk.Frame(frame_config, bg='white')
    selector_frame.pack(anchor='nw', padx=20, pady=(20, 10))

    label_canal = tk.Label(selector_frame, text='Canal de medición:', font=('Segoe UI', 10, 'bold'), fg='#25364a', bg='white')
    label_canal.pack(anchor='w')

    radio_frame = tk.Frame(selector_frame, bg='white')
    radio_frame.pack(anchor='w', pady=(5, 0))

    radio_a = tk.Radiobutton(radio_frame, text='Canal A', variable=canal_allan_var, value='A', font=('Segoe UI', 10), fg='#25364a', bg='white', selectcolor='white', activebackground='white', activeforeground='#2980f2')
    radio_a.pack(side='left', padx=(0, 15))
    radio_b = tk.Radiobutton(radio_frame, text='Canal B', variable=canal_allan_var, value='B', font=('Segoe UI', 10), fg='#25364a', bg='white', selectcolor='white', activebackground='white', activeforeground='#2980f2')
    radio_b.pack(side='left')

    # Por defecto, mostrar solo configuración
    frame_config.pack(fill='both', expand=True)
    frame_allan.pack_forget()

    # Función para cambiar de pestaña
    def mostrar_pestana(tab):
        if tab == 'config':
            frame_config.pack(fill='both', expand=True)
            frame_allan.pack_forget()
            btn_configuracion.config(bg='#2980f2', fg='white')
            btn_allan_dev.config(bg='#6c757d', fg='white')
        else:
            frame_allan.pack(fill='both', expand=True)
            frame_config.pack_forget()
            btn_configuracion.config(bg='#6c757d', fg='white')
            btn_allan_dev.config(bg='#2980f2', fg='white')

    # Asignar comandos a los botones
    btn_configuracion.config(command=lambda: mostrar_pestana('config'))
    btn_allan_dev.config(command=lambda: mostrar_pestana('allan'))

    # Campo para el número de muestras
    num_muestras_allan_var = tk.StringVar(value='10')

    muestras_frame = tk.Frame(frame_config, bg='white')
    muestras_frame.pack(anchor='nw', padx=20, pady=(0, 10))

    label_muestras = tk.Label(muestras_frame, text='Número de muestras:', font=('Segoe UI', 10, 'bold'), fg='#25364a', bg='white')
    label_muestras.pack(anchor='w')

    entry_muestras = tk.Entry(muestras_frame, textvariable=num_muestras_allan_var, font=('Segoe UI', 10), width=8, justify='left')
    entry_muestras.pack(anchor='w', pady=(5, 0))

    aviso_muestras = tk.Label(muestras_frame, text='', font=('Segoe UI', 9), fg='#e74c3c', bg='white')
    aviso_muestras.pack(anchor='w', pady=(2, 0))

    def validar_muestras(*args):
        try:
            valor = int(num_muestras_allan_var.get())
            if valor <= 1:
                aviso_muestras.config(text='Debe ser mayor a 1 para calcular el Allan Deviation')
            elif valor > 100:
                num_muestras_allan_var.set('100')
                aviso_muestras.config(text='')
            else:
                aviso_muestras.config(text='')
        except ValueError:
            if num_muestras_allan_var.get() != '':
                aviso_muestras.config(text='Introduce un número válido')
            else:
                aviso_muestras.config(text='')

    num_muestras_allan_var.trace('w', validar_muestras)

    # Campos para intervalo mínimo y máximo de captura
    intervalo_min_var = tk.StringVar(value='0.2')
    intervalo_max_var = tk.StringVar(value='1')

    intervalo_frame = tk.Frame(frame_config, bg='white')
    intervalo_frame.pack(anchor='nw', padx=20, pady=(0, 10))

    label_intervalo = tk.Label(intervalo_frame, text='Intervalo de captura (segundos):', font=('Segoe UI', 10, 'bold'), fg='#25364a', bg='white')
    label_intervalo.pack(anchor='w')

    sub_frame = tk.Frame(intervalo_frame, bg='white')
    sub_frame.pack(anchor='w', pady=(5, 0))

    label_min = tk.Label(sub_frame, text='Mínimo:', font=('Segoe UI', 10), fg='#25364a', bg='white')
    label_min.pack(side='left')
    entry_min = tk.Entry(sub_frame, textvariable=intervalo_min_var, font=('Segoe UI', 10), width=10, justify='left')
    entry_min.pack(side='left', padx=(5, 15))

    label_max = tk.Label(sub_frame, text='Máximo:', font=('Segoe UI', 10), fg='#25364a', bg='white')
    label_max.pack(side='left')
    entry_max = tk.Entry(sub_frame, textvariable=intervalo_max_var, font=('Segoe UI', 10), width=10, justify='left')
    entry_max.pack(side='left', padx=(5, 0))

    aviso_intervalo = tk.Label(intervalo_frame, text='', font=('Segoe UI', 9), fg='#e74c3c', bg='white')
    aviso_intervalo.pack(anchor='w', pady=(2, 0))

    def validar_intervalos(*args):
        try:
            min_val = float(intervalo_min_var.get())
        except ValueError:
            min_val = None
        try:
            max_val = float(intervalo_max_var.get())
        except ValueError:
            max_val = None
        aviso = ''
        # Validar rangos
        if min_val is not None:
            if min_val < 2e-8:
                intervalo_min_var.set('2e-8')
                min_val = 2e-8
            elif min_val > 1000:
                intervalo_min_var.set('1000')
                min_val = 1000
        if max_val is not None:
            if max_val < 2e-8:
                intervalo_max_var.set('2e-8')
                max_val = 2e-8
            elif max_val > 1000:
                intervalo_max_var.set('1000')
                max_val = 1000
        # Validar relación
        if min_val is not None and max_val is not None:
            if min_val >= max_val:
                aviso = 'El intervalo mínimo debe ser menor que el máximo.'
        aviso_intervalo.config(text=aviso)

    intervalo_min_var.trace('w', validar_intervalos)
    intervalo_max_var.trace('w', validar_intervalos)

    # Campo para pasos
    pasos_allan_var = tk.StringVar(value='5')

    pasos_frame = tk.Frame(frame_config, bg='white')
    pasos_frame.pack(anchor='nw', padx=20, pady=(0, 10))

    label_pasos = tk.Label(pasos_frame, text='Pasos:', font=('Segoe UI', 10, 'bold'), fg='#25364a', bg='white')
    label_pasos.pack(anchor='w')

    entry_pasos = tk.Entry(pasos_frame, textvariable=pasos_allan_var, font=('Segoe UI', 10), width=8, justify='left')
    entry_pasos.pack(anchor='w', pady=(5, 0))

    aviso_pasos = tk.Label(pasos_frame, text='', font=('Segoe UI', 9), fg='#e74c3c', bg='white')
    aviso_pasos.pack(anchor='w', pady=(2, 0))

    def validar_pasos(*args):
        try:
            valor = int(pasos_allan_var.get())
            if valor < 1:
                aviso_pasos.config(text='Debe ser al menos 1')
            else:
                aviso_pasos.config(text='')
        except ValueError:
            if pasos_allan_var.get() != '':
                aviso_pasos.config(text='Introduce un número válido')
            else:
                aviso_pasos.config(text='')

    pasos_allan_var.trace('w', validar_pasos)

    # Selector de Acoplamiento
    acoplamiento_allan_var = tk.StringVar(value='AC')
    acoplamiento_frame = tk.Frame(frame_config, bg='white')
    acoplamiento_frame.pack(anchor='nw', padx=20, pady=(0, 10))
    label_acoplamiento = tk.Label(acoplamiento_frame, text='Acoplamiento:', font=('Segoe UI', 10, 'bold'), fg='#25364a', bg='white')
    label_acoplamiento.pack(anchor='w')
    radio_acoplamiento_frame = tk.Frame(acoplamiento_frame, bg='white')
    radio_acoplamiento_frame.pack(anchor='w', pady=(5, 0))
    radio_ac = tk.Radiobutton(radio_acoplamiento_frame, text='AC', variable=acoplamiento_allan_var, value='AC', font=('Segoe UI', 10), fg='#25364a', bg='white', selectcolor='white', activebackground='white', activeforeground='#2980f2')
    radio_ac.pack(side='left', padx=(0, 15))
    radio_dc = tk.Radiobutton(radio_acoplamiento_frame, text='DC', variable=acoplamiento_allan_var, value='DC', font=('Segoe UI', 10), fg='#25364a', bg='white', selectcolor='white', activebackground='white', activeforeground='#2980f2')
    radio_dc.pack(side='left')

    # Selector de Impedancia
    impedancia_allan_var = tk.StringVar(value='50Ω')
    impedancia_frame = tk.Frame(frame_config, bg='white')
    impedancia_frame.pack(anchor='nw', padx=20, pady=(0, 10))
    label_impedancia = tk.Label(impedancia_frame, text='Impedancia:', font=('Segoe UI', 10, 'bold'), fg='#25364a', bg='white')
    label_impedancia.pack(anchor='w')
    radio_impedancia_frame = tk.Frame(impedancia_frame, bg='white')
    radio_impedancia_frame.pack(anchor='w', pady=(5, 0))
    radio_50 = tk.Radiobutton(radio_impedancia_frame, text='50Ω', variable=impedancia_allan_var, value='50Ω', font=('Segoe UI', 10), fg='#25364a', bg='white', selectcolor='white', activebackground='white', activeforeground='#2980f2')
    radio_50.pack(side='left', padx=(0, 15))
    radio_1M = tk.Radiobutton(radio_impedancia_frame, text='1MΩ', variable=impedancia_allan_var, value='1MΩ', font=('Segoe UI', 10), fg='#25364a', bg='white', selectcolor='white', activebackground='white', activeforeground='#2980f2')
    radio_1M.pack(side='left')

    # Selector de Atenuación
    atenuacion_allan_var = tk.StringVar(value='1x')
    atenuacion_frame = tk.Frame(frame_config, bg='white')
    atenuacion_frame.pack(anchor='nw', padx=20, pady=(0, 10))
    label_atenuacion = tk.Label(atenuacion_frame, text='Atenuación:', font=('Segoe UI', 10, 'bold'), fg='#25364a', bg='white')
    label_atenuacion.pack(anchor='w')
    radio_atenuacion_frame = tk.Frame(atenuacion_frame, bg='white')
    radio_atenuacion_frame.pack(anchor='w', pady=(5, 0))
    radio_1x = tk.Radiobutton(radio_atenuacion_frame, text='1x', variable=atenuacion_allan_var, value='1x', font=('Segoe UI', 10), fg='#25364a', bg='white', selectcolor='white', activebackground='white', activeforeground='#2980f2')
    radio_1x.pack(side='left', padx=(0, 15))
    radio_10x = tk.Radiobutton(radio_atenuacion_frame, text='10x', variable=atenuacion_allan_var, value='10x', font=('Segoe UI', 10), fg='#25364a', bg='white', selectcolor='white', activebackground='white', activeforeground='#2980f2')
    radio_10x.pack(side='left')

    # Selector de Nivel de Trigger
    trigger_mode_allan_var = tk.StringVar(value='automatico')
    trigger_value_allan_var = tk.StringVar(value='0')

    trigger_frame = tk.Frame(frame_config, bg='white')
    trigger_frame.pack(anchor='nw', padx=20, pady=(0, 10))
    label_trigger = tk.Label(trigger_frame, text='Nivel de Trigger:', font=('Segoe UI', 10, 'bold'), fg='#25364a', bg='white')
    label_trigger.pack(anchor='w')

    # Texto informativo de automático
    explicacion_trigger = tk.Label(trigger_frame, text='', font=('Segoe UI', 9), fg='#6c757d', bg='white')
    explicacion_trigger.pack(anchor='w', pady=(0, 3))

    def actualizar_explicacion_trigger(*args):
        canal_actual = canal_allan_var.get()
        if trigger_mode_allan_var.get() == 'automatico':
            if canal_actual == 'A':
                porcentaje = '70%'
            else:
                porcentaje = '30%'
            explicacion_trigger.config(text=f'(Automático: {porcentaje} del rango)')
        else:
            explicacion_trigger.config(text='')

    canal_allan_var.trace('w', actualizar_explicacion_trigger)
    trigger_mode_allan_var.trace('w', actualizar_explicacion_trigger)
    actualizar_explicacion_trigger()

    # Frame para los radio buttons de trigger
    radio_trigger_frame = tk.Frame(trigger_frame, bg='white')
    radio_trigger_frame.pack(anchor='w')
    radio_auto = tk.Radiobutton(radio_trigger_frame, text='Automático', variable=trigger_mode_allan_var, value='automatico', font=('Segoe UI', 10), fg='#25364a', bg='white', selectcolor='white', activebackground='white', activeforeground='#2980f2')
    radio_auto.pack(side='left', padx=(0, 15))
    radio_manual = tk.Radiobutton(radio_trigger_frame, text='Manual', variable=trigger_mode_allan_var, value='manual', font=('Segoe UI', 10), fg='#25364a', bg='white', selectcolor='white', activebackground='white', activeforeground='#2980f2')
    radio_manual.pack(side='left')

    # Frame para el valor manual del trigger
    trigger_value_frame = tk.Frame(trigger_frame, bg='white')
    trigger_value_frame.pack(anchor='w', pady=(5, 0))
    label_trigger_value = tk.Label(trigger_value_frame, text='Valor (-5 a 5 V):', font=('Segoe UI', 10), fg='#6c757d', bg='white')
    label_trigger_value.pack(side='left', padx=(0, 5))
    entry_trigger = tk.Entry(trigger_value_frame, textvariable=trigger_value_allan_var, font=('Segoe UI', 10), width=8, justify='left')
    entry_trigger.pack(side='left')

    aviso_trigger = tk.Label(trigger_value_frame, text='', font=('Segoe UI', 9), fg='#e74c3c', bg='white')
    aviso_trigger.pack(side='left', padx=(10, 0))

    def validar_entrada_trigger(*args):
        if trigger_mode_allan_var.get() != 'manual':
            aviso_trigger.config(text='')
            return
        try:
            texto_actual = trigger_value_allan_var.get()
            if texto_actual == '':
                aviso_trigger.config(text='')
                return
            valor = float(texto_actual)
            if valor < -5:
                trigger_value_allan_var.set('-5.0')
                aviso_trigger.config(text='')
            elif valor > 5:
                trigger_value_allan_var.set('5.0')
                aviso_trigger.config(text='')
            else:
                aviso_trigger.config(text='')
        except ValueError:
            aviso_trigger.config(text='Introduce un valor válido')

    trigger_value_allan_var.trace('w', validar_entrada_trigger)
    trigger_mode_allan_var.trace('w', lambda *a: validar_entrada_trigger())

    def actualizar_visibilidad_trigger_manual(*args):
        if trigger_mode_allan_var.get() == 'manual':
            trigger_value_frame.pack(anchor='w', pady=(5, 0))
        else:
            trigger_value_frame.pack_forget()
    trigger_mode_allan_var.trace('w', lambda *a: actualizar_visibilidad_trigger_manual())
    actualizar_visibilidad_trigger_manual()

    # Selector de Pendiente del Trigger
    trigger_slope_allan_var = tk.StringVar(value='POS')
    trigger_slope_frame = tk.Frame(frame_config, bg='white')
    trigger_slope_frame.pack(anchor='nw', padx=20, pady=(0, 10))
    label_trigger_slope = tk.Label(trigger_slope_frame, text='Pendiente del Trigger:', font=('Segoe UI', 10, 'bold'), fg='#25364a', bg='white')
    label_trigger_slope.pack(anchor='w')
    radio_trigger_slope_frame = tk.Frame(trigger_slope_frame, bg='white')
    radio_trigger_slope_frame.pack(anchor='w', pady=(5, 0))
    radio_pos = tk.Radiobutton(radio_trigger_slope_frame, text='Positiva', variable=trigger_slope_allan_var, value='POS', font=('Segoe UI', 10), fg='#25364a', bg='white', selectcolor='white', activebackground='white', activeforeground='#2980f2')
    radio_pos.pack(side='left', padx=(0, 15))
    radio_neg = tk.Radiobutton(radio_trigger_slope_frame, text='Negativa', variable=trigger_slope_allan_var, value='NEG', font=('Segoe UI', 10), fg='#25364a', bg='white', selectcolor='white', activebackground='white', activeforeground='#2980f2')
    radio_neg.pack(side='left')

    # Selector de Filtro Analógico pasa bajas
    filtro_analog_allan_var = tk.StringVar(value='False')
    filtro_analog_frame = tk.Frame(frame_config, bg='white')
    filtro_analog_frame.pack(anchor='nw', padx=20, pady=(0, 10))
    label_filtro_analog = tk.Label(filtro_analog_frame, text='Filtro Analógico pasa bajas:', font=('Segoe UI', 10, 'bold'), fg='#25364a', bg='white')
    label_filtro_analog.pack(anchor='w')
    radio_filtro_analog_frame = tk.Frame(filtro_analog_frame, bg='white')
    radio_filtro_analog_frame.pack(anchor='w', pady=(5, 0))
    radio_filtro_true = tk.Radiobutton(radio_filtro_analog_frame, text='True', variable=filtro_analog_allan_var, value='True', font=('Segoe UI', 10), fg='#25364a', bg='white', selectcolor='white', activebackground='white', activeforeground='#2980f2')
    radio_filtro_true.pack(side='left', padx=(0, 15))
    radio_filtro_false = tk.Radiobutton(radio_filtro_analog_frame, text='False', variable=filtro_analog_allan_var, value='False', font=('Segoe UI', 10), fg='#25364a', bg='white', selectcolor='white', activebackground='white', activeforeground='#2980f2')
    radio_filtro_false.pack(side='left')

    # Al crear los botones de pestaña, deshabilita el de Allan Deviation
    btn_allan_dev.config(state='disabled')

    # Función para habilitar el botón solo tras guardar
    def habilitar_allan_deviation():
        btn_allan_dev.config(state='normal')
    def deshabilitar_allan_deviation(*args):
        btn_allan_dev.config(state='disabled')

    # Añade trace a todos los StringVar relevantes para deshabilitar el botón si se cambia algo
    for var in [canal_allan_var, num_muestras_allan_var, intervalo_min_var, intervalo_max_var, pasos_allan_var,
                acoplamiento_allan_var, impedancia_allan_var, atenuacion_allan_var, trigger_mode_allan_var,
                trigger_value_allan_var, trigger_slope_allan_var, filtro_analog_allan_var]:
        var.trace('w', deshabilitar_allan_deviation)

    # Modifica la función guardar_configuracion_allan para habilitar el botón tras guardar
    def guardar_configuracion_allan():
        try:
            # Canal
            canal = canal_allan_var.get()
            # Número de muestras
            N_muestras = int(num_muestras_allan_var.get())
            # Intervalos
            intervalo_captura_min = float(intervalo_min_var.get())
            intervalo_captura_max = float(intervalo_max_var.get())
            # Pasos
            pasos = int(pasos_allan_var.get())
            # Acoplamiento
            acoplamiento = acoplamiento_allan_var.get()
            # Impedancia
            imp = impedancia_allan_var.get()
            if imp == '50Ω':
                impedancia = 'MIN'
            else:
                impedancia = 'MAX'
            # Atenuación
            atn = atenuacion_allan_var.get()
            if atn == '1x':
                atenuacion = '1'
            else:
                atenuacion = '10'
            # Trigger level
            if trigger_mode_allan_var.get() == 'automatico':
                trigger_level = None
            else:
                trigger_level = float(trigger_value_allan_var.get())
            # Trigger slope
            trigger_slope = trigger_slope_allan_var.get()
            # Filtro analógico
            filtro_Analog_PASSAbaja = True if filtro_analog_allan_var.get() == 'True' else False

            # Guardar en el diccionario global
            global allan_config
            allan_config = {
                'canal': canal,
                'N_muestras': N_muestras,
                'intervalo_captura_min': intervalo_captura_min,
                'intervalo_captura_max': intervalo_captura_max,
                'pasos': pasos,
                'acoplamiento': acoplamiento,
                'impedancia': impedancia,
                'atenuacion': atenuacion,
                'trigger_level': trigger_level,
                'trigger_slope': trigger_slope,
                'filtro_Analog_PASSAbaja': filtro_Analog_PASSAbaja
            }
            tk.messagebox.showinfo('Configuración guardada', 'La configuración de Allan Deviation ha sido guardada correctamente.')
            habilitar_allan_deviation()
        except Exception as e:
            tk.messagebox.showerror('Error', f'Error al guardar configuración: {e}')

    # Botón para guardar configuración
    btn_guardar_allan = tk.Button(frame_config, text='Guardar configuración', command=guardar_configuracion_allan, font=('Segoe UI', 10, 'bold'), bg='#2980f2', fg='white', relief='flat', padx=15, pady=5, cursor='hand2')
    btn_guardar_allan.pack(anchor='e', padx=20, pady=(20, 10))

# Función para mostrar la página de información del CNT-91
def mostrar_informacion_cnt91(widgets):
    frame_contenido = widgets['frame_contenido']
    
    # Limpiar el frame de contenido
    for widget in frame_contenido.winfo_children():
        widget.destroy()
    
    # Frame principal con padding
    main_frame = tk.Frame(frame_contenido, bg='#f6f7fa')
    main_frame.pack(fill='both', expand=True, padx=10, pady=5)
    
    # Título principal - Información CNT-91
    titulo_frame = tk.Frame(main_frame, bg='#f6f7fa')
    titulo_frame.pack(fill='x', pady=(0, 10))
    
    titulo = tk.Label(titulo_frame, text='Información CNT-91', 
                     font=('Segoe UI', 16, 'bold'), 
                     fg='#25364a', bg='#f6f7fa')
    titulo.pack(anchor='w')
    
    # Línea separadora
    separador = tk.Frame(titulo_frame, height=2, bg='#2980f2')
    separador.pack(fill='x', pady=(5, 0))
    
    # Frame de contenido con scroll
    contenido_frame = tk.Frame(main_frame, bg='white', relief='solid', bd=1)
    contenido_frame.pack(fill='both', expand=True, pady=(0, 5))
    
    # Canvas y scrollbar para contenido scrolleable
    canvas = tk.Canvas(contenido_frame, bg='white', highlightthickness=0, bd=0)
    scrollbar = ttk.Scrollbar(contenido_frame, orient="vertical", command=canvas.yview)
    # Frame centrador para el contenido
    centrador = tk.Frame(canvas, bg='white')
    scrollable_frame = tk.Frame(centrador, bg='white')
    scrollable_frame.pack(anchor='center', expand=True)

    # Centrar el contenido horizontalmente al redimensionar
    def resize_canvas(event):
        canvas_width = event.width
        centrador.config(width=canvas_width)
        canvas.itemconfig(window_id, width=canvas_width)
    canvas.bind('<Configure>', resize_canvas)

    centrador.pack(expand=True)
    window_id = canvas.create_window((0, 0), window=centrador, anchor="n")
    canvas.configure(yscrollcommand=scrollbar.set)

    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )

    # Empaquetar canvas y scrollbar con mejor espaciado
    canvas.pack(side="left", fill="both", expand=True, padx=(0, 5), pady=10)
    scrollbar.pack(side="right", fill="y", padx=(0, 10), pady=10)
    
    # Añadir margen izquierdo solo al contenido principal si se desea
    scrollable_frame.pack_configure(padx=10)
    
    # Contenido de la información
    contenido = get_info_cnt91_sections()
    
    # Crear el contenido con formato profesional
    for i, (titulo_seccion, texto) in enumerate(contenido):
        # Título de sección
        titulo_label = tk.Label(scrollable_frame, text=titulo_seccion,
                               font=('Segoe UI', 12, 'bold'),
                               fg='#25364a', bg='white',
                               anchor='w', justify='left')
        titulo_label.pack(fill='x', padx=15, pady=(20 if i == 0 else 15, 5))
        
        # Texto de la sección
        texto_label = tk.Label(scrollable_frame, text=texto,
                              font=('Segoe UI', 10),
                              fg='#2c3e50', bg='white',
                              anchor='nw', justify='left',
                              wraplength=800)
        texto_label.pack(fill='x', padx=15, pady=(0, 10))
        
        # Separador entre secciones (excepto la última)
        if i < len(contenido) - 1:
            separador_seccion = tk.Frame(scrollable_frame, height=1, bg='#e0e0e0')
            separador_seccion.pack(fill='x', padx=15, pady=5)
    
    # --- Sección de Documentación y Recursos ---
    recursos_frame = tk.Frame(scrollable_frame, bg='white')
    recursos_frame.pack(fill='x', padx=15, pady=(30, 10), anchor='w')

    # Título de la sección
    recursos_titulo = tk.Label(recursos_frame, text='Documentación y Recursos',
                               font=('Segoe UI', 13, 'bold'), fg='#25364a', bg='white', anchor='w')
    recursos_titulo.pack(anchor='w', pady=(0, 10))

    recursos = get_info_cnt91_resources()

    def abrir_url(url):
        import webbrowser
        webbrowser.open_new(url)

    for icono, texto, url in recursos:
        frame = tk.Frame(recursos_frame, bg='white')
        frame.pack(fill='x', pady=4, anchor='w')
        label_icon = tk.Label(frame, text=icono, font=('Segoe UI Symbol', 13), fg='#25364a', bg='white')
        label_icon.pack(side='left', padx=(0, 8))
        enlace = tk.Label(frame, text=texto, font=('Segoe UI', 11, 'bold'), fg='#25364a', bg='white', cursor='hand2')
        enlace.pack(side='left')
        # Efecto hover
        def on_enter(e, l=enlace):
            l.config(fg='#2980f2', underline=True)
        def on_leave(e, l=enlace):
            l.config(fg='#25364a', underline=False)
        enlace.bind('<Enter>', on_enter)
        enlace.bind('<Leave>', on_leave)
        enlace.bind('<Button-1>', lambda e, url=url: abrir_url(url))
    
    # Configurar el scroll para que funcione correctamente
    canvas.update_idletasks()
    canvas.configure(scrollregion=canvas.bbox("all"))

# Función para manejar la conexión al dispositivo
def conectar_dispositivo(widgets):
    global cnt_device
    entry_id = widgets['entry_id']
    estado_label = widgets['estado']
    btn_conectar = widgets['btn_conectar']
    btn_mediciones = widgets['btn_mediciones']
    btn_config = widgets['btn_config']
    address = entry_id.get().strip()
    # Si el campo está vacío, usar el valor por defecto
    if not address:
        address = DEFAULT_GPIB
        entry_id.delete(0, tk.END)
        entry_id.insert(0, DEFAULT_GPIB)
    # Cambiar estado a "Trying to connect" en naranja
    estado_label.config(text='Estado:  Trying to connect', fg='#e67e22')
    widgets['frame_superior'].update_idletasks()
    try:
        cnt_device = CNT.CNT_frequenciometro(address)
        # Si no hay excepción, conexión exitosa
        estado_label.config(text='Estado:  Conectado', fg='#27ae60')
        # Cambiar el botón a rojo con texto "Desconectar"
        btn_conectar.config(text='🔌  Desconectar', style='DangerSidebar.TButton')
        # Cambiar la función del botón para desconectar
        btn_conectar.config(command=lambda: desconectar_dispositivo(widgets))
        # Actualizar el estado del botón
        btn_conectar.estado_conectado = True
        # Habilitar botones que requieren conexión
        btn_mediciones.config(state='normal')
        btn_config.config(state='normal')
    except Exception as e:
        estado_label.config(text='Estado:  Desconectado', fg='#e74c3c')
        tk.messagebox.showerror('Error de conexión', 'No se logró conexión con el Dispositivo.\nRevise alimentación o instalación de drivers de comunicación.')
        btn_mediciones.config(state='disabled')

# Función para manejar la desconexión del dispositivo
def desconectar_dispositivo(widgets):
    global cnt_device
    estado_label = widgets['estado']
    btn_conectar = widgets['btn_conectar']
    btn_mediciones = widgets['btn_mediciones']
    btn_config = widgets['btn_config']
    
    try:
        if cnt_device is not None:
            # Cerrar la conexión usando la función que creamos
            cnt_device.cerrar_conexion()
            cnt_device = None
            # Cambiar estado a desconectado
            estado_label.config(text='Estado:  Desconectado', fg='#e74c3c')
            # Cambiar el botón de vuelta a azul con texto "Conectar"
            btn_conectar.config(text='🔌  Conectar', style='PrimarySidebar.TButton')
            # Cambiar la función del botón para conectar
            btn_conectar.config(command=lambda: conectar_dispositivo(widgets))
            # Actualizar el estado del botón
            btn_conectar.estado_conectado = False
            btn_mediciones.config(state='disabled')
            btn_config.config(state='disabled')
            print("Dispositivo desconectado correctamente.")
    except Exception as e:
        print(f"Error al desconectar: {e}")
        # Aún así, resetear el estado de la interfaz
        estado_label.config(text='Estado:  Desconectado', fg='#e74c3c')
        btn_conectar.config(text='🔌  Conectar', style='PrimarySidebar.TButton')
        btn_conectar.config(command=lambda: conectar_dispositivo(widgets))
        btn_conectar.estado_conectado = False
        btn_mediciones.config(state='disabled')
        btn_config.config(state='disabled')

# Función para manejar el cierre de la ventana
def on_closing(root, widgets):
    """Maneja el cierre correcto de la aplicación"""
    try:
        # Si hay un dispositivo conectado, desconectarlo
        if cnt_device is not None:
            print("Cerrando conexión con el dispositivo...")
            desconectar_dispositivo(widgets)
        
        # Si hay una medición en curso, finalizarla
        if 'datalogger_running' in globals() and datalogger_running.get():
            print("Finalizando medición en curso...")
            try:
                # Abortar medición continua
                cnt_device.abort_continuous_measurement()
            except:
                pass
        
        # Si hay un archivo Excel abierto, cerrarlo
        if 'ruta_archivo_excel' in globals() and ruta_archivo_excel is not None:
            print("Cerrando archivo Excel...")
            try:
                cnt_device.cerrar_archivo_excel()
            except:
                pass
        
        print("Cerrando aplicación...")
        
    except Exception as e:
        print(f"Error durante el cierre: {e}")
    
    finally:
        # Destruir la ventana principal
        root.destroy()
        # Salir del programa
        import sys
        sys.exit(0)

if __name__ == '__main__':
    # Crear la ventana principal de la aplicación
    root = tk.Tk()
    widgets = crear_layout_principal(root)

    # Configurar el protocolo de cierre de ventana
    root.protocol("WM_DELETE_WINDOW", lambda: on_closing(root, widgets))

    # Importar messagebox aquí para evitar problemas de importación circular
    import tkinter.messagebox
    tk.messagebox = tkinter.messagebox

    # Deshabilitar botones por defecto (requieren conexión)
    widgets['btn_mediciones'].config(state='disabled')
    widgets['btn_config'].config(state='disabled')

    # Asignar la función al botón de conectar
    widgets['btn_conectar'].config(command=lambda: conectar_dispositivo(widgets))

    # Asignar función al botón Allan Deviation vs tau
    widgets['btn_config'].config(command=lambda: mostrar_allan_deviation(widgets))

    # Asignar función al botón Mediciones
    widgets['btn_mediciones'].config(command=lambda: mostrar_menu_canal(widgets))

    # Asignar función al botón Información CNT-91
    widgets['btn_info'].config(command=lambda: mostrar_informacion_cnt91(widgets))

    # Iniciar el bucle principal de la interfaz gráfica (espera eventos del usuario)
    root.mainloop() 